from openpyxl import load_workbook
from pathlib import Path
from src.config import settings
from pathlib import Path
from src.schedule.service import schedule_service
from src.schedule.schemas import Teacher, TypeLesson, TypeDirection, Room, Discipline
from src.schemas import FullName


class ImportService:
    async def import_discipline(self, path: Path) -> None:
        wb = load_workbook(path)
        ws = wb.active
        for r in range(2, ws.max_column):
            await schedule_service.room_store.add(
                Discipline(
                    id=ws.cell(row=r, column=1).value,
                    name=ws.cell(row=r, column=2).value,
                    lecture_hours=ws.cell(row=r, column=3).value,
                    practice_hours=ws.cell(row=r, column=4).value
                )
            )

    async def import_room(self, path: Path) -> None:
        wb = load_workbook(path)
        ws = wb.active
        for r in range(2, ws.max_column):
            await schedule_service.room_store.add(
                Room(
                    id=ws.cell(row=r, column=1).value,
                    name=ws.cell(row=r, column=2).value,
                    profile=ws.cell(row=r, column=3).value
                )
            )

    async def import_teacher(self, path: Path) -> None:
        wb = load_workbook(path)
        ws = wb.active
        for r in range(2, ws.max_column):
            await schedule_service.teacher_store.add(
                Teacher(
                    id=ws.cell(row=r, column=1).value,
                    full_name=FullName(
                        surname=ws.cell(row=r, column=2).value,
                        name=ws.cell(row=r, column=3).value,
                        patronymic=ws.cell(row=r, column=4).value,
                    ),
                    position=str(ws.cell(row=r, column=5).value),
                    profile=str(ws.cell(row=r, column=6).value)
                )
            )

    async def import_type_lesson(self, path: Path) -> None:
        wb = load_workbook(path)
        ws = wb.active
        for r in range(2, ws.max_column):
            await schedule_service.type_lesson_store.add(
                TypeLesson(
                    id=ws.cell(row=r, column=1).value,
                    name=ws.cell(row=r, column=2).value
                )
            )

    async def import_type_direction(self, path: Path) -> None:
        wb = load_workbook(path)
        ws = wb.active
        for r in range(2, ws.max_column):
            await schedule_service.type_direction_store.add(
                TypeDirection(
                    id=ws.cell(row=r, column=1).value,
                    name=ws.cell(row=r, column=2).value,
                    full_name=ws.cell(row=r, column=3).value
                )
            )


import_service = ImportService()
