from openpyxl import Workbook
from pathlib import Path
from src.config import settings
from pathlib import Path
from src.constants import ExportConstants
from src.schedule.service import schedule_service
from src.schedule.schemas import Teacher, Direction, Discipline, Group
from typing import List
from src.schemas import FullName


class ExportService:
    async def export_teacher(self, dir_path: Path) -> Path:
        path: Path = dir_path / \
            f'export_teacher.{ExportConstants.FILE_EXTENSION}'
        teachers: List[Teacher] = await schedule_service.teacher_store.get_all()
        wb = Workbook()
        ws = wb.active
        ws.append([
            'ID',
            'Фамилия',
            'Имя',
            'Отчество',
            'Должность',
            'Профиль'
        ])
        for t in teachers:
            ws.append([
                t.id,
                t.full_name.surname,
                t.full_name.name,
                t.full_name.patronymic,
                t.position,
                t.profile
            ])
        wb.save(path)
        wb.close()
        return path

    async def export_direction(self, dir_path: Path) -> Path:
        path: Path = dir_path / \
            f'export_direction.{ExportConstants.FILE_EXTENSION}'

        directions: List[Direction] = await schedule_service.direction_store.get_all()
        wb = Workbook()
        ws = wb.active
        ws.append([
            'ID',
            'Название направления',
            'ID_SYS',
            'Тип направления',
            'Количество часов'
        ])
        for d in directions:
            ws.append([
                d.id,
                d.name,
                d.id_sys,
                d.type_direction.name,
                d.hours
            ])
        wb.save(path)
        wb.close()
        return path

    async def export_discipline(self, dir_path: Path) -> Path:
        path: Path = dir_path / \
            f'export_discipline.{ExportConstants.FILE_EXTENSION}'

        disciplines: List[Discipline] = await schedule_service.discipline_store.get_all()
        wb = Workbook()
        ws = wb.active
        ws.append([
            'ID',
            'Название дисциплины',
            'Количество лекционных часов',
            'Количество практических часов'
        ])
        for d in disciplines:
            ws.append([
                d.id,
                d.name,
                d.lecture_hours,
                d.practice_hours
            ])
        wb.save(path)
        wb.close()
        return path

    async def export_group(self, dir_path: Path) -> Path:
        path: Path = dir_path / \
            f'export_group.{ExportConstants.FILE_EXTENSION}'

        groups: List[Group] = await schedule_service.group_store.get_all()
        wb = Workbook()
        ws = wb.active
        ws.append([
            'ID',
            'Номер группы',
            'Название направления'
        ])
        for g in groups:
            ws.append([
                g.id,
                g.number_group,
                g.direction.name
            ])
        wb.save(path)
        wb.close()

        return path

    async def export_schedule(
            self,
            group_id: int,
            schedule_list_id: int,
            dir_path: Path
    ) -> Path:

        path: Path = dir_path / \
            f'export_schedule.{ExportConstants.FILE_EXTENSION}'
        schedules = await schedule_service.schedule_store.get_by_schedule_list_id(schedule_list_id)
        group = await schedule_service.group_store.get(group_id)
        wb = Workbook()
        ws = wb.active
        ws.append([
            f'Группа №{group.number_group}'
        ])
        ws.append(
            [
                'Дата',
                'Время начала',
                'Время окончания',
                'Название дисциплины',
                'Преподаватель',
                'Кабинет',
                'Тип занятия'
            ]
        )

        for s in schedules:
            if group_id in [g.id for g in s.flow.groups]:
                teachers = ''
                for s_t in s.schedule_teachers:
                    if s_t.change:
                        t = s_t.change.teacher.full_name
                    else:
                        t = s_t.teacher.full_name
                    teachers += f'{t.surname} {t.name} {t.patronymic}; '
                ws.append(
                    [
                        s.date_,
                        s.time_start,
                        s.time_end,
                        s.discipline.name,
                        teachers,
                        s.room.name,
                        s.type_lesson.name
                    ]
                )
        wb.save(path)
        wb.close()

        return path
    

    async def export_schedule_for_all_flows(
            self,
            schedule_list_id,
            dir_path: Path
    ):
        path: Path = dir_path / \
            f'export_schedule.{ExportConstants.FILE_EXTENSION}'
        flows = await schedule_service.flow_store.get_all()
        wb = Workbook()
        ws = wb.active
        for f in flows:
            ws.title = f'Поток {f.name}'

            schedules = await schedule_service.schedule_store.get_by_flow_id(flow_id=f.id, schedule_list_id=schedule_list_id)

            for s in schedules:
                teachers = ''
                for s_t in s.schedule_teachers:
                    if s_t.change:
                        t = s_t.change.teacher.full_name
                    else:
                        t = s_t.teacher.full_name
                    teachers += f'{t.surname} {t.name} {t.patronymic}; '
                ws.append(
                    [
                        s.date_,
                        s.time_start,
                        s.time_end,
                        s.discipline.name,
                        teachers,
                        s.room.name,
                        s.type_lesson.name
                    ]
                )
            ws = wb.create_sheet()
        wb.save(path)
        wb.close()

        return path


export_service = ExportService()
